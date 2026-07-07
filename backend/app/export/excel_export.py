"""Excel export mirroring the school's own manual review workflow: they
build the timetable in Excel first to sanity-check it before re-entering
into Vigilo by hand. One sheet per class, rows = periods, columns = days.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.db.models.activity import Activity, ActivityLeg
from app.db.models.period import DayOfWeek, PeriodDefinition
from app.db.models.subject import Subject
from app.db.models.teacher import Teacher
from app.db.models.timetable import GeneratedTimetable, TimetableSlot
from app.db.models.trinn_class import ClassGroup, SchoolClass, Trinn
from app.solver.grid import WEEK_ORDER, build_tick_grid

_DAY_LABELS = {
    DayOfWeek.MON: "Mandag",
    DayOfWeek.TUE: "Tirsdag",
    DayOfWeek.WED: "Onsdag",
    DayOfWeek.THU: "Torsdag",
    DayOfWeek.FRI: "Fredag",
}

_HEADER_FILL = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")
_HEADER_FONT = Font(bold=True)


def export_generated_timetable(db: Session, generated_timetable_id: int) -> Workbook:
    generated = db.get(GeneratedTimetable, generated_timetable_id)
    if generated is None:
        raise ValueError(f"GeneratedTimetable {generated_timetable_id} not found")
    school_year_id = generated.school_year_id

    periods = db.scalars(
        select(PeriodDefinition).where(PeriodDefinition.school_year_id == school_year_id)
    ).all()
    grid = build_tick_grid(list(periods))
    max_period_by_day: dict[DayOfWeek, int] = {}
    for p in periods:
        max_period_by_day[p.day_of_week] = max(max_period_by_day.get(p.day_of_week, 0), p.period_number)

    slots = db.scalars(
        select(TimetableSlot).where(TimetableSlot.generated_timetable_id == generated_timetable_id)
    ).all()

    activities = db.scalars(
        select(Activity)
        .where(Activity.school_year_id == school_year_id)
        .options(selectinload(Activity.legs).selectinload(ActivityLeg.leg_teachers))
    ).all()
    activities_by_id = {a.id: a for a in activities}

    subjects_by_id = {s.id: s for s in db.scalars(select(Subject).where(Subject.school_year_id == school_year_id)).all()}
    teachers_by_id = {t.id: t for t in db.scalars(select(Teacher)).all()}

    classes = db.scalars(
        select(SchoolClass)
        .join(Trinn, SchoolClass.trinn_id == Trinn.id)
        .where(Trinn.school_year_id == school_year_id)
        .order_by(Trinn.level, SchoolClass.name)
    ).all()
    class_groups = db.scalars(
        select(ClassGroup).where(ClassGroup.school_class_id.in_([c.id for c in classes]))
    ).all()
    class_group_ids_by_class: dict[int, set[int]] = {}
    for g in class_groups:
        class_group_ids_by_class.setdefault(g.school_class_id, set()).add(g.id)

    # cell_text[class_id][(day, period)] -> list of leg-text strings
    cell_text: dict[int, dict[tuple[DayOfWeek, int], list[str]]] = {c.id: {} for c in classes}

    for slot in slots:
        activity = activities_by_id.get(slot.activity_id)
        if activity is None:
            continue
        periods_touched = grid.touched_periods(slot.day_of_week, slot.start_tick, slot.duration_ticks)
        for class_ in classes:
            group_ids = class_group_ids_by_class.get(class_.id, set())
            matching_legs = [leg for leg in activity.legs if leg.class_group_id in group_ids]
            if not matching_legs:
                continue
            for leg in matching_legs:
                subject = subjects_by_id.get(leg.subject_id)
                subject_code = subject.short_code if subject else "?"
                teacher_initials = [
                    teachers_by_id[lt.teacher_id].initials
                    for lt in leg.leg_teachers
                    if lt.teacher_id in teachers_by_id
                ]
                text = f"{subject_code} ({'/'.join(teacher_initials)})" if teacher_initials else subject_code
                for period in periods_touched:
                    key = (slot.day_of_week, period)
                    cell_text[class_.id].setdefault(key, [])
                    if text not in cell_text[class_.id][key]:
                        cell_text[class_.id][key].append(text)

    workbook = Workbook()
    workbook.remove(workbook.active)

    for class_ in classes:
        _write_class_sheet(workbook, class_, max_period_by_day, cell_text[class_.id])

    return workbook


def _write_class_sheet(
    workbook: Workbook,
    school_class: SchoolClass,
    max_period_by_day: dict[DayOfWeek, int],
    cell_text: dict[tuple[DayOfWeek, int], list[str]],
) -> None:
    sheet: Worksheet = workbook.create_sheet(title=school_class.name)
    sheet.cell(row=1, column=1, value="Periode").font = _HEADER_FONT

    days_present = [d for d in WEEK_ORDER if d in max_period_by_day]
    for col, day in enumerate(days_present, start=2):
        cell = sheet.cell(row=1, column=col, value=_DAY_LABELS[day])
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    max_periods = max(max_period_by_day.values()) if max_period_by_day else 0
    for period in range(1, max_periods + 1):
        row = period + 1
        sheet.cell(row=row, column=1, value=period).font = _HEADER_FONT
        for col, day in enumerate(days_present, start=2):
            if period > max_period_by_day.get(day, 0):
                continue  # e.g. Tuesday has no period 5/6
            texts = cell_text.get((day, period), [])
            cell = sheet.cell(row=row, column=col, value="\n".join(texts) if texts else "")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    sheet.column_dimensions["A"].width = 10
    for col in range(2, 2 + len(days_present)):
        sheet.column_dimensions[chr(ord("A") + col - 1)].width = 22
