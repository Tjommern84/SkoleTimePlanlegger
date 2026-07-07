from openpyxl import load_workbook
from io import BytesIO

from app.export.excel_export import export_generated_timetable
from app.solver.solve_service import solve_school_year
from tests.fixtures.school_example_data import seed_school_example_data


def test_export_creates_one_sheet_per_class_with_expected_content(db_session):
    result = seed_school_example_data(db_session)
    school_year_id = result["school_year"].id

    solve_result = solve_school_year(db_session, school_year_id, time_limit_seconds=30)
    assert solve_result.status in ("OPTIMAL", "FEASIBLE")

    from datetime import UTC, datetime

    from app.db.models.timetable import GeneratedTimetable, TimetableSlot

    generated = GeneratedTimetable(
        school_year_id=school_year_id,
        created_at=datetime.now(UTC),
        solver_status=solve_result.status,
        objective_value=None,
        is_active=True,
    )
    db_session.add(generated)
    db_session.flush()
    for p in solve_result.placements:
        db_session.add(
            TimetableSlot(
                generated_timetable_id=generated.id,
                activity_id=p.activity_id,
                occurrence_index=p.occurrence_index,
                day_of_week=p.day_of_week,
                start_tick=p.start_tick,
                duration_ticks=p.duration_ticks,
            )
        )
    db_session.commit()

    workbook = export_generated_timetable(db_session, generated.id)
    sheet_names = set(workbook.sheetnames)
    assert sheet_names == {"8A", "8B", "8C", "9A", "9B", "9C", "10A", "10B", "10C"}

    # Round-trip through bytes to make sure it's a valid xlsx file.
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    reloaded = load_workbook(buffer)
    sheet_8a = reloaded["8A"]
    assert sheet_8a["A1"].value == "Periode"
    assert sheet_8a["B1"].value == "Mandag"

    # 8A has Norsk co-taught (GB/EB) on 3 of its sessions -- at least one
    # cell across the sheet should mention Norsk's short code "NO".
    found_norsk = any(
        cell.value and "NO" in str(cell.value)
        for row in sheet_8a.iter_rows(min_row=2)
        for cell in row
    )
    assert found_norsk

    # 9A has a split Mat&Helse/Naturfag session -- some cell should show
    # both subject codes stacked (from the two legs of that activity).
    sheet_9a = reloaded["9A"]
    found_split = any(
        cell.value and "MH" in str(cell.value) and "NAT" in str(cell.value)
        for row in sheet_9a.iter_rows(min_row=2)
        for cell in row
    )
    assert found_split
